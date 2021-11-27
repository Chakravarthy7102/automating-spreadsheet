import openpyxl as xl
from openpyxl.chart import BarChart, Series, Reference
from pathlib import Path


# using the workbook


def update_spreedSheet(filename):
    workBook = xl.load_workbook(filename)
    sheet = workBook['Sheet1']
    cell = sheet.cell(1, 1)

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        updated_value = cell.value*1000
        updated_cell = sheet.cell(row, 4)
        updated_cell.value = updated_value

    values = Reference(sheet, min_row=2, max_row=sheet.max_row,
                       min_col=4, max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e3')
    workBook.save(filename)
    print("done!!")


path = Path()
xls = path.glob('*.xlsx')

for file in xls:

    update_spreedSheet(file)


# filename = 'transactions.xlsx'
# Update_spreedSheet(filename)
